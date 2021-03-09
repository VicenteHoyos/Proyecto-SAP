
import tkinter as tk
from tkinter import ttk
import pandas as pd

root = tk.Tk()

sample = {"File Name":[f"file_{i}" for i in range(5)],
          'Sheet Name': [f"sheet_{i}" for i in range(5)],
          'Number Of Rows': [f"row_{i}" for i in range(5)],
          'Number Of Columns': [f"col_{i}" for i in range(5)]
          }
df = pd.DataFrame(sample)
cols = list(df.columns)
print( cols)

tree = ttk.Treeview(root)
tree.pack()
tree["columns"] = cols

for i in cols:
    tree.column(i, anchor="w")
    tree.heading(i, text=i, anchor='w')
"""
for index, row in df.iterrows():
    tree.insert("",0,text=index,values=list(row))"""

root.mainloop()
'''from tkinter import *
from tkinter import ttk  
from tkinter import filedialog  # apertura de archivo
from openpyxl import load_workbook  # 

root = Tk()
root.title('Recuperación de datos de Excel')
root.geometry('1500x700')

"""Configuration des onglets"""
my_notebook = ttk.Notebook(root)
my_notebook.pack()

"""Fonction permettant d'ouvrir un fichier dans l'ordinateur"""
my_program = filedialog.askopenfilename()  # ouverture de la boîte de dialogue

"""Chargement de la feuille active d'Excel"""
wb = load_workbook(my_program)


class OpenProgram(Frame):
    """Clase utilizada para recuperar datos del archivo de Excel y mostrarlos en Python con el módulo Tkinter
    """

    def __init__(self, master, number, start_cell, end_cell, row_py, column_py):
        super().__init__(master)
        self.pack()
        self.widgets(number, start_cell, end_cell, row_py, column_py)

    def widgets(self, number, start_cell, end_cell, row_py, column_py):

        """Assignation de la feuille d'Excel (feuille = worksheet = ws)"""
        ws = wb[wb.sheetnames[number]]  # number = Número de pestaña de archivo de Excel

        """Saisie de la plage de cellules des données à récupérer du fichier Excel"""
        range_excel = ws[start_cell:end_cell]

        """Assignation d'une liste"""
        my_list = []

        """Insertion des données Excel dans une liste"""
        for cell in range_excel:
            for x in cell:
                my_list.append(x.value)  # ajout des données de la feuille Excel dans une liste

        print(my_list)


        """Création d'un cadre principal qui s'étend sur toute la fenêtre"""
        main_frame = Frame(self)
        main_frame.pack()

        """Création d'un canvas situé à gauche du cadre principal"""
        my_canvas = Canvas(main_frame, width=1460, height=660)
        my_canvas.pack(side=LEFT)

        """Ajout d'une barre de défilement en bas du canvas sur toute sa longueur"""
        my_scrollbar_bottom = ttk.Scrollbar(main_frame, orient=VERTICAL, command=my_canvas.yview)
        my_scrollbar_bottom.pack(side=RIGHT, fill=Y)

        """Configuration du canvas"""
        my_canvas.config(yscrollcommand=my_scrollbar_bottom.set)
        my_canvas.bind('<Configure>', lambda e: my_canvas.config(scrollregion=my_canvas.bbox('all')))

        """Création d'un autre cadre à l'intérieur du canvas"""
        second_frame = Frame(my_canvas)

        """Ajout de ce nouveau cadre à la fenêtre du canvas"""
        my_canvas.create_window((0, 0), window=second_frame, anchor=NW)

        """Assignation de variables"""
        row_excel = row_py  # nombre de lignes +1
        column_excel = column_py  # nombre de colonnes +1
        nb = 0


        listbox = Listbox(root)
        """Ajout des champs de saisis"""
        for i in range(1, row_excel):  # rows
            for j in range(1, column_excel):  # columns
                cell = Entry(second_frame, justify='center', font='arial 6', width=15)
                cell.grid(row=i, column=j, ipady=5)
                if my_list[nb] is None:
                    cell.insert(0, '')
                    listbox.insert('end', "")
                else:
                    cell.insert(0,
                                my_list[nb])  # insertion des données de la liste 'my_list' dans les champs de saisies

                    listbox.insert('end', my_list[nb])            
                nb += 1  # incrémentation +1 des données de la liste 'my_list'
                #print(nb)

        listbox.place(x=615, y =40)
        
"""Configuration des cadres"""
frame0 = Frame(my_notebook)
#frame1 = Frame(my_notebook)

"""Ajout des onglets"""
my_notebook.add(frame0, text='Actif du bilan')
#my_notebook.add(frame1, text='Passif du bilan')

open0 = OpenProgram(frame0, 0, 'A1', 'N51', 25, 15)
#open1 = OpenProgram(frame1, 1, 'A1', 'D20', 21, 5)
root.mainloop()'''